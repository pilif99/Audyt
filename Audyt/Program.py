import openpyxl
from openpyxl import Workbook
from Indeksy import *

class Program:

    indeksy = openpyxl.load_workbook(r"C:\Users\Filip\Desktop\zestawienie tytulow opisow.xlsx").active
    plik1 = openpyxl.load_workbook(r"C:\Users\Filip\Desktop\Zeszyt1.xlsx")
    plik = plik1.active
    lista_indeksow = set()
    
    def __init__(self):

        self.stworzenie_listy_indeksow()
        self.stworzenie_indeksow()
        self.wypisanie_danych()

    def stworzenie_listy_indeksow(self):
        wiersz = 2
        while self.indeksy.cell(wiersz,1).value != None:
        #self.indeksy.cell(wiersz,1).value != None:
            self.lista_indeksow.add(self.indeksy.cell(wiersz, column=1).value)
            wiersz += 1
        list(self.lista_indeksow)

    def stworzenie_indeksow(self):

        for i in self.lista_indeksow:
            for j in ['DE', 'ES', 'IT']:
                x = list(i)
                for k in x:
                    if k == '-':
                        x.remove(k)
                globals()[''.join(x)] = IndeksSKU(i, j)

    def wypisanie_danych(self):
        wiersz = 1
        for i in ParentSKU.lista_parentow:
            ## parent
            for k in range(len(ParentSKU.lista_parentow[i].informacje_o_parencie()) - 1):
                x = list(i)
                for l in range(len(x) - 2):
                    del x[0]
                self.plik.cell(wiersz, 1).value = ParentSKU.lista_parentow[i].parent_sku
                self.plik.cell(wiersz, 2).value = 1
                self.plik.cell(wiersz, 3).value = ''.join(x)
                self.plik.cell(wiersz, 4).value = ParentSKU.lista_parentow[i].informacje_o_parencie()[k + 1]
                self.plik.cell(wiersz, 5).value = getattr(ParentSKU.lista_parentow[i], ParentSKU.lista_parentow[i].informacje_o_parencie()[k + 1])
                wiersz += 1

            ## indeksy
            x = list(i)
            for l in range(len(x) - 2):
                del x[0]
            a = getattr(ParentSKU.lista_parentow[i], 'lista')
            for j in range(len(a)):
                for m in range(len(a[j].informacje(''.join(x))) - 1):
                    self.plik.cell(wiersz, 1).value = a[j].indeksSKU
                    self.plik.cell(wiersz, 2).value = 1
                    self.plik.cell(wiersz, 3).value = ''.join(x)
                    self.plik.cell(wiersz, 4).value = a[j].informacje(''.join(x))[m + 1]
                    self.plik.cell(wiersz, 5).value = getattr(a[j], a[j].informacje(''.join(x))[m + 1])
                    wiersz += 1

        self.plik1.save(r"C:\Users\Filip\Desktop\Zeszyt1.xlsx")

a = Program()