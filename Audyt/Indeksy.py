
import openpyxl

class Kolory:

    kolory = openpyxl.load_workbook(r"C:\Users\Filip\Desktop\kolory.xlsx").active

    def __init__(self):
     
        wiersz = 1
        kolumna = 1
        while self.kolory.cell(wiersz, kolumna).value != None:
            while self.kolory.cell(wiersz, kolumna).value != None:
                if wiersz == 1:
                    a = list(self.kolory.cell(wiersz, kolumna).value)
                    for i in a:
                        if i == ' ':
                            a.remove(i)
                    b = ''.join(a)
                    globals()[b] = []
                else:
                    globals()[b].append(self.kolory.cell(wiersz, kolumna).value)
                wiersz += 1
            wiersz = 1
            kolumna += 1

class Indeksy:
    
    def parent_sku(self, sku):
  
        x = list(sku)
        a = x.index('_')
        b = x.index('_', x.index('_') + 1)
        c = []
        for i in range(a):
            c.append(x[i])
        if x[a - 4] + x[a - 3] + x[a - 2] + x[a - 1] == 'LADY':
            for i in range(5):
                del c[len(c) - 1]
            c.append('_L')
        elif x[b + 1] == 'D':
            c.append('_L')
        else:
            c.append('_P')
        return ''.join(c)

class ParentSKU(Indeksy):

    lista_parentow = {}

    def __init__(self, sku, jezyk):

        # zmienne
        self.parent_sku = self.parent_sku(sku)
        self.feed_product_type = self.feed_product_type()
        self.plec = self.plec()
        self.brand_name = self.brand_name(jezyk)
        self.manufacturer = self.manufacturer(jezyk)
        self.country_of_origin = self.country_of_origin(jezyk)
        self.parent_child = self.parent_child(jezyk)
        self.department_name = self.department_name(jezyk)
        self.target_gender = self.target_gender(jezyk)
        self.outer_material_type = self.outer_material_type(jezyk)
        self.material_type = self.material_type(jezyk)
        self.model_name = self.model_name(jezyk)
        self.variation_theme = self.variation_theme(jezyk)

        # lista indeksów

        x = list(self.parent_sku)
        for k in x:
            if k == '-':
                x.remove(k)
        a = ''.join(x)

        globals()['lista_' + a + jezyk] = []
        self.lista = globals()['lista_' + a + jezyk]

    # funkcje potrzebne do stworzenia parenta
    
    def variation_theme(self, jezyk):
        if self.feed_product_type == 'coat' and jezyk == 'NL':
            return 'Sizename-Kleurnaam'
        elif jezyk == 'ES':
            return 'color-size'
        else:
            return 'SizeColor'

    def feed_product_type(self):

        x = list(self.parent_sku)
        a = x.index('-', 3) - 3
        for i in range(3):
            del x[0]
        while len(x) > a:
            del x[a]
        bsku = ''.join(x)
        if bsku == 'TJ' or bsku == 'LJ' or bsku == 'JRY':
            return 'coat'
        elif bsku == 'TP' or bsku == 'LP' or bsku == 'JP':
            return 'pants'
        elif bsku == 'B' or bsku == 'BOT':
            return 'boot'
        elif bsku == 'GLV':
            return 'sportactivityglove'
        elif bsku == 'LS1':
            return 'suit'
        else:
            print(self.parent_sku)

    def plec(self):
        if self.feed_product_type == 'boot':
            return 'unisex'
        else:
            x = list(self.parent_sku)
            if x[len(x) - 1] == 'L':
                return 'female'
            else:
                return 'male'

    def brand_name(self, jezyk):
        a = list(self.parent_sku)
        asku = a[0] + a[1]
        if asku == 'RH':
            return 'REBELHORN'
        elif asku == 'OZ':
            return 'OZONE'
        elif asku == 'BR':
            return 'BROGER'
        else:
            return print('Nie rozpoznano marki')

    def manufacturer(self, jezyk):
        
        return self.brand_name.title()

    def country_of_origin(self, jezyk):
        if jezyk == 'ES':
            return 'Pakistán'
        else:
            return 'Pakistan'

    def parent_child(self, jezyk):

        return 'parent'

    def department_name(self, jezyk):
        Kobiety = {'DE': 'Damen', 'ES': 'Mujer', 'IT': 'Donna', 'NL': 'Vrouwen'}
        Mezczyzni = {'DE': 'Herren', 'ES': 'Hombre', 'IT': 'Uomo', 'NL': 'Mens'}
        if self.plec == 'unisex':
            return 'unisex'
        elif self.plec == 'male':
            return Mezczyzni.get(jezyk)
        else:
            return Kobiety.get(jezyk)

    def target_gender(self, jezyk):
        x = {'Herren': 'Männlich', 'Damen': 'Weiblich', 'unisex': 'unisex', 'Hombre': 'Masculino', 'Mujer': 'Femenino', 'Uomo': 'Maschio', 'Donna': 'Femmina', 'Vrouwen': 'Vrouwelijk', 'Mens': 'Mannelijk'}
        return x.get(self.department_name)

    def outer_material_type(self, jezyk):
        syntetyk = {'DE': 'Synthetisch', 'ES': 'sintético', 'IT': 'sintetico', 'NL': 'Synthetisch'}
        skora = {'DE': 'Leder', 'ES': 'cuero', 'IT': 'pelle', 'NL': 'Leder'}
        jeans = {'DE': 'Jeans', 'ES': 'denim', 'IT': 'denim', 'NL': 'Denim'}
        if self.feed_product_type == 'sportactivityglove':
            return skora.get(jezyk)
        elif self.feed_product_type == 'boot':
            return syntetyk.get(jezyk)
        else:
            x = list(self.parent_sku)
            if x[3] == 'T' and x[len(x) - 3] == '/' or x[3] == 'J':
                return jeans.get(jezyk)
            elif x[3] == 'L':
                return skora.get(jezyk)
            else:
                return syntetyk.get(jezyk)

    def material_type(self, jezyk):
        skora = {'DE': 'Leder', 'ES': 'cuero', 'IT': 'pelle', 'NL': 'Leder'}
        return skora.get(jezyk)

    def model_name(self, jezyk):

        x = list(self.parent_sku)
        a = x.index('-', 3)
        for i in range(a + 1):
            del x[0]
        if x[len(x) - 6] + x[len(x) - 5] + x[len(x) - 4] + x[len(x) - 3] == 'LADY':
            b = len(x)
            for j in range(5):
                del x[b - 7]
        for k in range(2):
            del x[len(x) - 1]
        
        if self.feed_product_type == 'coat':
            x.append(' COAT')
        elif self.feed_product_type == 'pants':
            x.append(' PANTS')
        elif self.feed_product_type == 'sportactivityglove':
            x.append(' GLOVES')
        elif self.feed_product_type == 'boot':
            x.append(' BOOTS')
        elif self.feed_product_type == 'suit':
            x.append(' SUIT')
        if self.plec == 'female':
            x.append(' LADY')
        return ''.join(x)

    #informacje o parencie

    def informacje_o_parencie(self):

        return ['parent_sku', 'brand_name', 'manufacturer', 'parent_child', 'variation_theme', 'model_name']

class IndeksSKU(Indeksy):

    def __init__(self, sku, jezyk):

        self.parent_sku = self.parent_sku(sku)

        # sprawdzenie czy parent istnieje

        x = list(self.parent_sku)
        for k in x:
            if k == '-':
                x.remove(k)
        a = ''.join(x)
        if not ''.join(a + jezyk) in ParentSKU.lista_parentow:

            ParentSKU.lista_parentow[a + jezyk] = ParentSKU(sku, jezyk)
        
        #zmienne
        kolor = Kolory()
        self.indeksSKU = sku
        self.feed_product_type = ParentSKU.lista_parentow[a + jezyk].feed_product_type
        self.plec = ParentSKU.lista_parentow[a + jezyk].plec
        self.color_name = self.color_name(jezyk)
        self.color_map = self.color_map(jezyk)
        self.model = self.model(jezyk)
        self.part_number = self.part_number(jezyk)
        self.parent_child = self.parent_child(jezyk)
        self.brand_name = ParentSKU.lista_parentow[a + jezyk].brand_name
        self.manufacturer = ParentSKU.lista_parentow[a + jezyk].manufacturer
        self.country_of_origin = ParentSKU.lista_parentow[a + jezyk].country_of_origin
        self.department_name = ParentSKU.lista_parentow[a + jezyk].department_name
        self.target_gender = ParentSKU.lista_parentow[a + jezyk].target_gender
        self.outer_material_type = ParentSKU.lista_parentow[a + jezyk].outer_material_type
        self.material_type = ParentSKU.lista_parentow[a + jezyk].material_type
        self.model_name = ParentSKU.lista_parentow[a + jezyk].model_name
        self.variation_theme = ParentSKU.lista_parentow[a + jezyk].variation_theme
        self.size_name = self.size_name(jezyk)
        self.size_map = self.size_map(jezyk)
        self.footwear_size = self.footwear_size(jezyk)

        # dodanie tego indeksu do listy parenta

        globals()['lista_' + a + jezyk].append(self)
    
    # funkcje potrzebne do stworzenia indeksu 

    def color_name(self, jezyk):

        a = SKU.index(self.indeksSKU)
        if jezyk == 'DE':
            return Kolorniemiecki[a]
        elif jezyk == 'ES':
            return Kolorhiszpański[a]
        if jezyk == 'IT':
            return Kolorwłoski[a]

    def color_map(self, jezyk):
        a = SKU.index(self.indeksSKU)
        if jezyk == 'DE':
            return KolormapaDE[a]
        elif jezyk == 'ES':
            return KolormapaES[a]
        if jezyk == 'IT':
            return KolormapaIT[a]

    def part_number(self, jezyk):

        return self.indeksSKU

    def model(self, jezyk):

        return self.indeksSKU

    def parent_child(self, jezyk):

        return 'child'

    def size_name(self, jezyk):
        x = list(self.indeksSKU)
        a = x.index('_', x.index('_') + 1) + 1
        if x[a] == 'D':
            a += 1
        for i in range(a):
            del x[0]
        if len(x) > 2:
            if x[2] == '/':
                x[2] = 'W / '
                x.append('L')
        return ''.join(x)

    def size_map(self, jezyk):

        return self.size_name

    def footwear_size(self, jezyk):

        return self.size_name

    ## informacje

    def informacje(self, jezyk):

        informacje = []
        informacje.extend(['indeksSKU', 'brand_name', 'manufacturer', 'part_number', 'parent_child', 'variation_theme', 'parent_sku'])
        if self.feed_product_type != 'boot' and jezyk == 'NL':
            informacje.append('department_name1')
        elif self.feed_product_type != 'boot':
            informacje.append('department_name')
        informacje.extend(['model_name', 'color_name', 'color_map'])
        if self.feed_product_type == 'sportactivityglove':
            informacje.append('material_type')
        else:
            informacje.append('outer_material_type')
        informacje.append('country_of_origin')
        if self.feed_product_type == 'boot':
            informacje.append('footwear_size')
        else:
            informacje.extend(['size_name', 'size_map'])
        informacje.append('model')
        if self.feed_product_type == 'coat' or self.feed_product_type == 'boot':
            informacje.append('target_gender')
        return informacje